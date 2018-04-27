#CTRL+ALT+N to Run the python script
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import requests
import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook
from openpyxl import load_workbook
import re

#x = urllib.request.urlopen('https://pythonprogramming.net')

try:

    #Stop remember me from showing up
    chrome_options = Options()
    chrome_options.add_experimental_option('prefs', {
        'credentials_enable_service': False,
        'profile': {
            'password_manager_enabled': False
        }
    })

    #Open the chrome driver, and navigate to the page
    driver = webdriver.Chrome("newchromedriver/chromedriver.exe",chrome_options=chrome_options)
    

    #Load original Excel file - and select the needed sheet - get maximum row number
    Result_File = load_workbook('source_file/IPOs.xlsx')
    Result_sheet = Result_File.worksheets[0]
    Max_Row_Count = Result_sheet.max_row

    #Count not found records  on site
    Not_Found_Tickers = 0
    
    #Filing type
    Filing_Type = "424B4"

    #Program Name
    Program_Name = "directed share program"

    for i in range(920, Max_Row_Count):

        driver.get("https://www.sec.gov/edgar/searchedgar/companysearch.html")

        #Get the ticker and EPS date from data sheet
        print(i)  

        #Save a final results sheet
        if(i%5 == 0):
            Result_File.save('results_file/FINAL.xlsx')  

        Ticker_Symbol = Result_sheet.cell(row=i, column=3).value

        #Locate the search box on top of the page
        Ticker_Symbol_Search = driver.find_element_by_name("CIK")

        #send ticker symbol value to the search box
        Ticker_Symbol_Search.send_keys(str(Ticker_Symbol))
        
        #Press enter on the search box
        Ticker_Symbol_Search.send_keys(Keys.ENTER)

        try:
            time.sleep(2)  
            #Find the Filing type type search box
            Filing_Type_Search = driver.find_element_by_name('type')

            #send filing type to search box
            Filing_Type_Search.send_keys(Filing_Type)

            #press enter on the search box
            Filing_Type_Search.send_keys(Keys.ENTER)

            time.sleep(2)           

            #Finding number of rows in SEC Table
            results_table = driver.find_element_by_class_name("tableFile2")
            results_rows = len(results_table.find_elements_by_tag_name("tr"))
            #IPO Date from the sheet
            IPO_Date = Result_sheet.cell(row=i, column=1).value
            #array of all results
            date_calculation = []
            #Find the required date for the IPO
            for j in range (2, results_rows+1):
                date_search_result = datetime.strptime(driver.find_element_by_xpath('//*[@id="seriesDiv"]/table/tbody/tr['+str(j)+']/td[4]').text, '%Y-%m-%d')
                date_minus_search_result = IPO_Date - date_search_result
                date_minus_search_result = date_minus_search_result.days
                date_calculation.append(date_minus_search_result) 
                
            #Find the closest cell to the required
            closest_date = min(date_calculation, key=abs)
            #Get the index of the array that had the closest date
            for z in range(0, len(date_calculation)):
                if closest_date == date_calculation[z]:
                    closest_date = z
                    break

            #Find the row that has the closest date
            Document_cell = driver.find_element_by_xpath('//*[@id="seriesDiv"]/table/tbody/tr['+str(closest_date+2)+']/td[2]').click()
            
            time.sleep(2)

            #Grab the first link for the 424B4 and click it
            Documents_Link = driver.find_element_by_xpath('//*[@id="formDiv"]/div/table/tbody/tr[2]/td[3]/a').click()

    

            time.sleep(2)
            #get text from the page body
            Body = driver.find_element_by_xpath("/html/body")
            Body_To_Text = Body.text
            Body_To_Text = Body_To_Text.lower()
            

            Program_Text = re.findall(r'([^.]*'+Program_Name+'[^.]*)', Body_To_Text)
        
            Program_Name = "directed shares"

            Program_Text = Program_Text + re.findall(r'([^.]*'+Program_Name+'[^.]*)', Body_To_Text)

            Program_Name = "directed unit program"

            Program_Text = Program_Text + re.findall(r'([^.]*'+Program_Name+'[^.]*)', Body_To_Text)

            if not Program_Text:
                Result_sheet.cell(row=i, column=13).value = 0
                Result_sheet.cell(row=i, column=14).value = 'No Shares Program'
                Result_sheet.cell(row=i, column=15).value = 'No Shares Program'

            else:
                Result_sheet.cell(row=i, column=13).value = 1
                Result_sheet.cell(row=i, column=14).value = ''
                Result_sheet.cell(row=i, column=15).value = '.'.join(Program_Text)

        except Exception as e:
            Result_sheet.cell(row=i, column=13).value = 'No SEC Record'
            Result_sheet.cell(row=i, column=14).value = 'No SEC Record'
            Result_sheet.cell(row=i, column=15).value = 'No SEC Record'
            Not_Found_Tickers += 1
            print('not found '+ str(Not_Found_Tickers))
            
    
    #Save a final results sheet
    Result_File.save('results_file/FINAL.xlsx')   

    #close the chrome page
    driver.close()

except Exception as e:
    print(str(e))